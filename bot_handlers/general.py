from __future__ import annotations

import logging
import os
from collections import Counter
from datetime import datetime, timedelta

from money import format_money
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from booking_service import consume_booking_start_payload
from booking_validation import normalize_phone, parse_working_hours, slot_overlaps
from time_utils import get_salon_now
from .states import AdminAvailabilityForm, AdminBookingsByDateForm, AdminEditBookingForm, AdminRescheduleBookingForm, ManualBookingForm, SearchBookingForm

from .base import (
    Command,
    CommandObject,
    F,
    FSInputFile,
    Router,
    database,
    escape,
    getenv,
    keyboards,
    salon_config,
    types,
)

router = Router()
logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="F4C7D9")
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="FAE6EE")
THIN_BORDER = Border(
    left=Side(style="thin", color="E7BCCB"),
    right=Side(style="thin", color="E7BCCB"),
    top=Side(style="thin", color="E7BCCB"),
    bottom=Side(style="thin", color="E7BCCB"),
)
BOOKINGS_PAGE_SIZE = 10
ADMIN_SERVICE_PAGE_SIZE = 12
CLIENTS_PAGE_SIZE = 10

SOURCE_LABELS = {
    "telegram": "Telegram",
    "whatsapp": "WhatsApp",
    "instagram": "Instagram",
    "phone": "Звонок",
    "offline": "Офлайн",
    "manual": "Вручную",
}

MENU_ESCAPE_TEXTS = {
    "👤 Главное меню",
    "👤 Меню клиента",
    "⚙️ Панель управления",
    "🗓 Все записи",
    "🗓 На сегодня",
    "📅 По дате",
    "📊 Статистика",
    "👥 Клиенты",
    "📃 Excel",
    "🔎 Поиск",
    "🕒 Свободные окна",
    "➕ Внести запись",
    "⚙️ Услуги",
    "📁 Категории",
    "🗓 Онлайн-запись",
    "💎 Услуги и цены",
    "🗓 Актуальные записи",
    "🕘 История",
}


def _status_label(status: str) -> str:
    return {
        "scheduled": "Активна",
        "completed": "Выполнена",
        "no_show": "Не пришел",
        "cancelled": "Отменена",
    }.get(status, status)


def _safe_parse_date(value: str) -> datetime | None:
    try:
        return datetime.strptime(value, "%d.%m.%Y")
    except ValueError:
        return None


def _source_label(source: str | None) -> str:
    return SOURCE_LABELS.get((source or "").strip(), source or "—")


def _format_admin_date_label(target_date: datetime.date) -> str:
    weekdays = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]
    return f"{target_date.strftime('%d.%m')} · {weekdays[target_date.weekday()]}"


def _format_iso_to_date(value: str | None) -> str:
    if not value:
        return "—"
    try:
        return datetime.fromisoformat(value).strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return value


async def _notify_client_about_admin_booking_action(
    bot,
    *,
    user_id: int | None,
    title: str,
    lines: list[str],
) -> None:
    if not user_id:
        return

    text = title + "\n\n" + "\n".join(lines)
    try:
        await bot.send_message(int(user_id), text, parse_mode="HTML")
    except Exception:
        logger.exception("Failed to notify Telegram client about admin booking action", extra={"user_id": user_id})


async def _build_admin_date_options(duration: int) -> list[tuple[str, str]]:
    salon_now = get_salon_now()
    booking_window = max(int(salon_config.get("booking_window", 7) or 7), 1)
    working_days = salon_config.get("working_days", [1, 2, 3, 4, 5, 6, 0])
    blacklisted_dates = set(salon_config.get("blacklisted_dates", []))
    options: list[tuple[str, str]] = []

    for offset in range(booking_window):
        target_date = salon_now.date() + timedelta(days=offset)
        date_value = target_date.strftime("%d.%m.%Y")
        js_weekday = (target_date.weekday() + 1) % 7
        if js_weekday not in working_days or date_value in blacklisted_dates:
            continue

        times = await _build_admin_time_options(date_value, duration)
        if times:
            options.append((_format_admin_date_label(target_date), date_value))

    return options


async def _build_admin_time_options(date_value: str, duration: int) -> list[str]:
    start_mins, end_mins = parse_working_hours(salon_config.get("working_hours", "10:00-20:00"))
    interval = int(salon_config.get("schedule_interval", 30) or 30)
    if interval <= 0:
        interval = 30

    busy_slots = await database.get_busy_slots_by_date(date_value)
    salon_now = get_salon_now()
    salon_today = salon_now.strftime("%d.%m.%Y")
    current_salon_mins = salon_now.hour * 60 + salon_now.minute if date_value == salon_today else -1

    available_times: list[str] = []
    duration_minutes = int(duration or 60)
    for slot_start in range(start_mins, end_mins - duration_minutes + 1, interval):
        if slot_start <= current_salon_mins:
            continue

        is_busy = False
        for busy in busy_slots:
            try:
                busy_h, busy_m = map(int, busy["time"].split(":"))
            except (KeyError, ValueError, AttributeError):
                continue
            busy_start = busy_h * 60 + busy_m
            busy_duration = int(busy.get("duration") or 60)
            if slot_overlaps(slot_start, duration_minutes, busy_start, busy_duration):
                is_busy = True
                break

        if not is_busy:
            available_times.append(f"{slot_start // 60:02d}:{slot_start % 60:02d}")

    return available_times


def _find_service_by_id(services: list[dict], service_id: int | None) -> dict | None:
    if service_id is None:
        return None
    for service in services:
        if int(service.get("id") or 0) == int(service_id):
            return service
    return None


def _fit_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        first_real_cell = next((cell for cell in column_cells if not isinstance(cell, MergedCell)), None)
        if first_real_cell is None:
            continue
        column_letter = get_column_letter(first_real_cell.column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)


def _style_table_header(row) -> None:
    for cell in row:
        cell.font = Font(bold=True, color="6B2B43")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_data_rows(ws, start_row: int, end_row: int, center_columns: set[int] | None = None) -> None:
    center_columns = center_columns or set()
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.column in center_columns:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def _build_bookings_workbook(file_path: str, bookings: list[tuple[str, str, str, str, int | None, str]]) -> None:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Сводка"
    bookings_ws = workbook.create_sheet("Записи")
    daily_ws = workbook.create_sheet("По дням")

    salon_name = salon_config.get("salon_name", "Салон")
    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M")
    parsed_dates = [parsed for _, _, date, _, _, _ in bookings if (parsed := _safe_parse_date(date))]
    min_date = min(parsed_dates).strftime("%d.%m.%Y") if parsed_dates else "-"
    max_date = max(parsed_dates).strftime("%d.%m.%Y") if parsed_dates else "-"
    daily_counts = Counter(date for _, _, date, _, _, _ in bookings)

    summary_rows = [
        ("Отчет", f"Записи салона «{salon_name}»"),
        ("Сформирован", generated_at),
        ("Всего записей", len(bookings)),
        ("Период данных", f"{min_date} - {max_date}"),
        ("Уникальных телефонов", len({phone for _, phone, _, _, _, _ in bookings if phone})),
        ("Сумма по записям", format_money(sum(int(price or 0) for _, _, _, _, price, _ in bookings))),
    ]

    summary_ws["A1"] = "Экспорт записей"
    summary_ws["A1"].font = Font(size=16, bold=True, color="6B2B43")
    summary_ws["A1"].alignment = Alignment(vertical="center")
    summary_ws.merge_cells("A1:B1")

    for idx, (label, value) in enumerate(summary_rows, start=3):
        label_cell = summary_ws[f"A{idx}"]
        value_cell = summary_ws[f"B{idx}"]
        label_cell.value = label
        value_cell.value = value
        label_cell.font = Font(bold=True, color="6B2B43")
        label_cell.fill = SUBHEADER_FILL
        label_cell.border = THIN_BORDER
        value_cell.border = THIN_BORDER

    _fit_columns(summary_ws)

    bookings_ws.append(["№", "Клиент / услуга", "Телефон", "Дата", "Время", "Цена", "Статус"])
    for idx, (name, phone, date, time, price, status) in enumerate(bookings, start=1):
        bookings_ws.append([idx, name, phone, date, time, int(price or 0), _status_label(status)])
    _style_table_header(bookings_ws[1])
    if bookings_ws.max_row > 1:
        _style_data_rows(bookings_ws, 2, bookings_ws.max_row, center_columns={1, 4, 5, 6, 7})
    bookings_ws.freeze_panes = "A2"
    bookings_ws.auto_filter.ref = bookings_ws.dimensions
    for cell in bookings_ws["F"][1:]:
        cell.number_format = f'#,##0 "{salon_config.get("currency_symbol", "₸")}"'
    _fit_columns(bookings_ws)

    daily_ws.append(["Дата", "Количество записей"])
    for date, count in sorted(daily_counts.items(), key=lambda item: (_safe_parse_date(item[0]) or datetime.max)):
        daily_ws.append([date, count])
    _style_table_header(daily_ws[1])
    if daily_ws.max_row > 1:
        _style_data_rows(daily_ws, 2, daily_ws.max_row, center_columns={2})
    daily_ws.freeze_panes = "A2"
    daily_ws.auto_filter.ref = daily_ws.dimensions
    _fit_columns(daily_ws)

    workbook.save(file_path)


def _build_all_bookings_export_workbook(file_path: str, bookings: list[tuple]) -> None:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Сводка"
    bookings_ws = workbook.create_sheet("Все записи")

    salon_name = salon_config.get("salon_name", "Салон")
    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M")
    parsed_dates = [parsed for *_, date, _time, _duration, _price, _status, _source, _notes, _created_by_admin, _created_at in bookings if (parsed := _safe_parse_date(date))]
    min_date = min(parsed_dates).strftime("%d.%m.%Y") if parsed_dates else "-"
    max_date = max(parsed_dates).strftime("%d.%m.%Y") if parsed_dates else "-"

    summary_rows = [
        ("Отчет", f"Все записи салона «{salon_name}»"),
        ("Сформирован", generated_at),
        ("Всего записей", len(bookings)),
        ("Период данных", f"{min_date} - {max_date}"),
        ("Выполненных", sum(1 for booking in bookings if booking[8] == "completed")),
        ("Отмененных", sum(1 for booking in bookings if booking[8] == "cancelled")),
        ("Не пришли", sum(1 for booking in bookings if booking[8] == "no_show")),
    ]

    summary_ws["A1"] = "Экспорт всех записей"
    summary_ws["A1"].font = Font(size=16, bold=True, color="6B2B43")
    summary_ws.merge_cells("A1:B1")
    for idx, (label, value) in enumerate(summary_rows, start=3):
        summary_ws[f"A{idx}"] = label
        summary_ws[f"B{idx}"] = value
        summary_ws[f"A{idx}"].font = Font(bold=True, color="6B2B43")
        summary_ws[f"A{idx}"].fill = SUBHEADER_FILL
        summary_ws[f"A{idx}"].border = THIN_BORDER
        summary_ws[f"B{idx}"].border = THIN_BORDER
    _fit_columns(summary_ws)

    bookings_ws.append(
        ["№", "Клиент", "Телефон", "Услуга", "Дата", "Время", "Длит.", "Цена", "Статус", "Источник", "Комментарий", "Создано"]
    )
    for idx, (_id, name, phone, service_name, date, time, duration, price, status, source, notes, created_by_admin, created_at) in enumerate(bookings, start=1):
        bookings_ws.append(
            [
                idx,
                name,
                phone,
                service_name or "—",
                date,
                time,
                int(duration or 0),
                int(price or 0),
                _status_label(status),
                _source_label(source),
                notes or "",
                _format_iso_to_date(created_at),
            ]
        )
    _style_table_header(bookings_ws[1])
    if bookings_ws.max_row > 1:
        _style_data_rows(bookings_ws, 2, bookings_ws.max_row, center_columns={1, 5, 6, 7, 8, 9})
    bookings_ws.freeze_panes = "A2"
    bookings_ws.auto_filter.ref = bookings_ws.dimensions
    for cell in bookings_ws["H"][1:]:
        cell.number_format = f'#,##0 "{salon_config.get("currency_symbol", "₸")}"'
    _fit_columns(bookings_ws)
    workbook.save(file_path)


def _build_completed_services_workbook(file_path: str, bookings: list[tuple]) -> None:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Сводка"
    completed_ws = workbook.create_sheet("Выполненные")

    salon_name = salon_config.get("salon_name", "Салон")
    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M")
    total_revenue = sum(int(booking[7] or 0) for booking in bookings)

    summary_rows = [
        ("Отчет", f"Выполненные услуги салона «{salon_name}»"),
        ("Сформирован", generated_at),
        ("Всего выполнено", len(bookings)),
        ("Сумма", format_money(total_revenue)),
    ]

    summary_ws["A1"] = "Отчет по выполненным услугам"
    summary_ws["A1"].font = Font(size=16, bold=True, color="6B2B43")
    summary_ws.merge_cells("A1:B1")
    for idx, (label, value) in enumerate(summary_rows, start=3):
        summary_ws[f"A{idx}"] = label
        summary_ws[f"B{idx}"] = value
        summary_ws[f"A{idx}"].font = Font(bold=True, color="6B2B43")
        summary_ws[f"A{idx}"].fill = SUBHEADER_FILL
        summary_ws[f"A{idx}"].border = THIN_BORDER
        summary_ws[f"B{idx}"].border = THIN_BORDER
    _fit_columns(summary_ws)

    completed_ws.append(["№", "Клиент", "Телефон", "Услуга", "Дата", "Время", "Длит.", "Цена", "Источник", "Комментарий", "Выполнено"])
    for idx, (_id, name, phone, service_name, date, time, duration, price, _status, source, notes, _created_by_admin, completed_at) in enumerate(bookings, start=1):
        completed_ws.append(
            [
                idx,
                name,
                phone,
                service_name or "—",
                date,
                time,
                int(duration or 0),
                int(price or 0),
                _source_label(source),
                notes or "",
                _format_iso_to_date(completed_at),
            ]
        )
    _style_table_header(completed_ws[1])
    if completed_ws.max_row > 1:
        _style_data_rows(completed_ws, 2, completed_ws.max_row, center_columns={1, 5, 6, 7, 8})
    completed_ws.freeze_panes = "A2"
    completed_ws.auto_filter.ref = completed_ws.dimensions
    for cell in completed_ws["H"][1:]:
        cell.number_format = f'#,##0 "{salon_config.get("currency_symbol", "₸")}"'
    _fit_columns(completed_ws)
    workbook.save(file_path)


def _build_clients_workbook(file_path: str, clients: list[tuple]) -> None:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Сводка"
    clients_ws = workbook.create_sheet("Клиенты")

    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M")
    total_completed_revenue = sum(int(client[7] or 0) for client in clients)
    summary_rows = [
        ("Сформирован", generated_at),
        ("Клиентов в базе", len(clients)),
        ("Всего выполненных визитов", sum(int(client[4] or 0) for client in clients)),
        ("Сумма по выполненным", format_money(total_completed_revenue)),
    ]

    summary_ws["A1"] = "Клиентская база"
    summary_ws["A1"].font = Font(size=16, bold=True, color="6B2B43")
    summary_ws.merge_cells("A1:B1")
    for idx, (label, value) in enumerate(summary_rows, start=3):
        summary_ws[f"A{idx}"] = label
        summary_ws[f"B{idx}"] = value
        summary_ws[f"A{idx}"].font = Font(bold=True, color="6B2B43")
        summary_ws[f"A{idx}"].fill = SUBHEADER_FILL
        summary_ws[f"A{idx}"].border = THIN_BORDER
        summary_ws[f"B{idx}"].border = THIN_BORDER
    _fit_columns(summary_ws)

    clients_ws.append(["№", "Клиент", "Телефон", "Всего записей", "Выполнено", "Отменено", "Не пришел", "Сумма", "Последняя запись", "Последнее создание"])
    for idx, (_phone_key, name, phone, total, completed, cancelled, no_show, revenue, last_date_iso, last_created_at) in enumerate(clients, start=1):
        last_visit = "—"
        if last_date_iso:
            try:
                last_visit = datetime.fromisoformat(last_date_iso).strftime("%d.%m.%Y")
            except ValueError:
                last_visit = last_date_iso
        clients_ws.append(
            [
                idx,
                name,
                phone or "—",
                int(total or 0),
                int(completed or 0),
                int(cancelled or 0),
                int(no_show or 0),
                int(revenue or 0),
                last_visit,
                _format_iso_to_date(last_created_at),
            ]
        )
    _style_table_header(clients_ws[1])
    if clients_ws.max_row > 1:
        _style_data_rows(clients_ws, 2, clients_ws.max_row, center_columns={1, 4, 5, 6, 7, 8})
    clients_ws.freeze_panes = "A2"
    clients_ws.auto_filter.ref = clients_ws.dimensions
    for cell in clients_ws["H"][1:]:
        cell.number_format = f'#,##0 "{salon_config.get("currency_symbol", "₸")}"'
    _fit_columns(clients_ws)
    workbook.save(file_path)


def _paginate(items, page: int, page_size: int = BOOKINGS_PAGE_SIZE):
    total = len(items)
    total_pages = max((total - 1) // page_size + 1, 1)
    page = max(0, min(page, total_pages - 1))
    start = page * page_size
    end = start + page_size
    return items[start:end], page, total_pages


def _get_last_page(items, page_size: int = BOOKINGS_PAGE_SIZE) -> int:
    total = len(items)
    total_pages = max((total - 1) // page_size + 1, 1)
    return total_pages - 1


def _render_booking_page(bookings, title: str, page: int):
    page_items, page, total_pages = _paginate(bookings, page)
    lines = [f"🗓 <b>{title}</b>", f"<i>Страница {page + 1} из {total_pages}</i>", ""]
    for idx, (_booking_id, name, phone, date, time, price, status) in enumerate(page_items, start=1 + page * BOOKINGS_PAGE_SIZE):
        safe_name = escape(name)
        safe_phone = escape(phone)
        safe_date = escape(date)
        safe_time = escape(time)
        status_badge = {
            "scheduled": "🟢 Активна",
            "completed": "✅ Выполнена",
            "no_show": "🟠 Не пришел",
            "cancelled": "❌ Отменена",
        }.get(status, escape(_status_label(status)))
        lines.append(
            f"┌ <b>Запись #{idx}</b>\n"
            f"├ <b>Когда:</b> {safe_date} в {safe_time}\n"
            f"├ <b>Клиент:</b> {safe_name}\n"
            f"├ <b>Телефон:</b> {safe_phone}\n"
            f"├ <b>Статус:</b> {status_badge}\n"
            f"└ <b>Сумма:</b> {escape(format_money(price))}\n"
        )
    if not page_items:
        lines.append("Записей пока нет.")
    return "\n".join(lines), page_items, page, total_pages


async def _show_booking_list(message_or_callback, *, context: str, page: int = 0):
    await database.sync_completed_bookings()
    if context == "today":
        today_str = datetime.now().strftime("%d.%m.%Y")
        bookings = await database.get_bookings_by_date_detailed(today_str)
        title = f"Записи на сегодня ({today_str})"
    else:
        bookings = await database.get_all_bookings_detailed()
        title = "Все записи"

    text, page_items, page, total_pages = _render_booking_page(bookings, title, page)
    markup = keyboards.get_admin_booking_page_keyboard(page_items, context, page, total_pages)

    if isinstance(message_or_callback, types.CallbackQuery):
        await message_or_callback.message.edit_text(text, parse_mode="HTML", reply_markup=markup)
    else:
        await message_or_callback.answer(text, parse_mode="HTML", reply_markup=markup)


async def send_client_home(message: types.Message, *, text: str, is_admin: bool) -> None:
    await message.answer(
        text,
        parse_mode="HTML",
        reply_markup=keyboards.get_booking_launch_keyboard(),
    )
    await message.answer(
        "👤 <b>Личный кабинет</b>\n\n"
        "Здесь можно посмотреть свои записи, историю и быстро вернуться к оформлению новой записи.",
        parse_mode="HTML",
        reply_markup=keyboards.get_main_menu(is_admin=is_admin),
    )


@router.message(Command("start"))
async def start_handler(message: types.Message, command: CommandObject | None = None, state=None):
    admin_id = getenv("ADMIN_ID")
    is_admin = bool(admin_id and str(message.from_user.id) == admin_id)
    if state is not None:
        await state.clear()

    if is_admin:
        await message.answer(
            "⚙️ <b>Панель администратора</b>\n\nВыберите нужный раздел ниже.",
            parse_mode="HTML",
            reply_markup=keyboards.admin_menu,
        )
        return

    payload = (command.args or "").strip() if command else ""
    if payload.startswith("booking_"):
        booking_id = consume_booking_start_payload(payload, user_id=message.from_user.id)
        if booking_id:
            booking = await database.get_booking_admin_details(int(booking_id))
            if booking:
                if len(booking) == 13:
                    _booking_id, user_id, name, phone, date, time, status, duration, service_name, price, source, notes, created_by_admin = booking
                    created_at = None
                else:
                    _booking_id, user_id, name, phone, date, time, status, duration, service_name, price, source, notes, created_by_admin, created_at = booking
                if int(user_id or 0) == int(message.from_user.id):
                    await message.answer(
                        (
                            "📅 <b>Запись оформлена</b>\n\n"
                            f"<b>Услуга:</b> {escape(service_name or '—')}\n"
                            f"<b>Дата:</b> {escape(date)}\n"
                            f"<b>Время:</b> {escape(time)}\n"
                            f"<b>Телефон:</b> {escape(phone or '—')}\n"
                            "Ждём вас в выбранное время."
                        ),
                        parse_mode="HTML",
                    )
                    return

    welcome_text = salon_config.get("welcome_text", "Привет! Выберите нужное действие:")
    await send_client_home(message, text=welcome_text, is_admin=is_admin)

@router.message(F.text == "👤 Главное меню")
@router.message(F.text == "👤 Меню клиента")
async def client_menu_handler(message: types.Message, state=None):
    admin_id = getenv("ADMIN_ID")
    is_admin = bool(admin_id and str(message.from_user.id) == admin_id)
    if not is_admin:
        return
    if state is not None:
        await state.clear()
    await send_client_home(
        message,
        text=(
            "📅 <b>Онлайн-запись</b>\n\n"
            "Выберите удобное время и оформите запись онлайн."
        ),
        is_admin=is_admin,
    )



@router.message(Command("admin"))
@router.message(F.text == "⚙️ Панель управления")
async def admin_handler(message: types.Message, state=None):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return
    if state is not None:
        await state.clear()
    await message.answer(
        "⚙️ <b>Панель администратора</b>\n\nВыберите нужный раздел ниже.",
        parse_mode="HTML",
        reply_markup=keyboards.admin_menu,
    )


@router.callback_query(F.data == "back_to_admin_menu")
async def back_to_admin_menu_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    await state.clear()
    try:
        await callback.message.delete()
    except Exception:
        pass
    await callback.message.answer(
        "⚙️ <b>Панель администратора</b>\n\nВыберите нужный раздел ниже.",
        parse_mode="HTML",
        reply_markup=keyboards.admin_menu,
    )


@router.callback_query(F.data == "cancel_admin_action")
async def cancel_admin_action_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    await state.clear()
    await callback.message.delete()


async def _send_excel_export(
    message: types.Message,
    *,
    file_path: str,
    caption: str,
    build_fn,
    rows,
):
    if not rows:
        await message.answer("📃 <b>Для выбранной выгрузки пока нет данных</b>", parse_mode="HTML")
        return

    build_fn(file_path, rows)
    try:
        await message.answer_document(FSInputFile(file_path), caption=caption, parse_mode="HTML")
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


@router.message(F.text == "📃 Excel")
async def export_excel_menu_handler(message: types.Message):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return

    await message.answer(
        "📃 <b>Excel-выгрузки</b>\n\nВыберите нужный формат отчета.",
        parse_mode="HTML",
        reply_markup=keyboards.get_excel_exports_keyboard(),
    )


@router.callback_query(F.data == "excel_export_all")
async def export_all_excel_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    rows = await database.get_all_bookings_export()
    await _send_excel_export(
        callback.message,
        file_path="bookings_all_export.xlsx",
        caption="📋 <b>Все записи</b>",
        build_fn=_build_all_bookings_export_workbook,
        rows=rows,
    )


@router.callback_query(F.data == "excel_export_completed")
async def export_completed_excel_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    rows = await database.get_completed_bookings_export()
    await _send_excel_export(
        callback.message,
        file_path="bookings_completed_export.xlsx",
        caption="✅ <b>Выполненные услуги</b>",
        build_fn=_build_completed_services_workbook,
        rows=rows,
    )


@router.callback_query(F.data == "excel_export_clients")
async def export_clients_excel_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    rows = await database.get_client_base_export()
    await _send_excel_export(
        callback.message,
        file_path="clients_export.xlsx",
        caption="👥 <b>Клиентская база</b>",
        build_fn=_build_clients_workbook,
        rows=rows,
    )


@router.message(F.text == "👥 Клиенты")
async def clients_menu_handler(message: types.Message):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return

    clients = await database.get_client_base_export()
    text, page_items, page, total_pages = _render_clients_page(clients, 0)
    await message.answer(
        text,
        parse_mode="HTML",
        reply_markup=_build_clients_keyboard(page_items, page, total_pages),
    )


@router.callback_query(F.data.startswith("clients_page|"))
async def clients_page_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    _, page_token = callback.data.split("|", 1)
    if page_token == "noop":
        return

    clients = await database.get_client_base_export()
    text, page_items, page, total_pages = _render_clients_page(clients, int(page_token))
    await callback.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=_build_clients_keyboard(page_items, page, total_pages),
    )


@router.callback_query(F.data.startswith("clients_open|"))
async def clients_open_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    _, page_str, index_str = callback.data.split("|", 2)
    page = int(page_str)
    index = int(index_str)
    clients = await database.get_client_base_export()
    page_items, page, total_pages = _paginate(clients, page, page_size=CLIENTS_PAGE_SIZE)
    if index < 0 or index >= len(page_items):
        await callback.answer("Клиент не найден", show_alert=True)
        return

    client = page_items[index]
    await callback.message.edit_text(
        _format_client_card(client),
        parse_mode="HTML",
        reply_markup=_build_clients_keyboard(page_items, page, total_pages),
    )


@router.message(F.text == "➕ Внести запись")
async def manual_booking_start_handler(message: types.Message, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return

    services = await database.get_all_services()
    if not services:
        await message.answer("⚙️ Сначала добавьте хотя бы одну услугу.", parse_mode="HTML")
        return

    await state.clear()
    await state.set_state(ManualBookingForm.service_id)
    await state.update_data(service_page=0)
    await message.answer(
        "➕ <b>Ручная запись</b>\n\nВыберите услугу для новой записи.",
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_service_picker_keyboard(services, prefix="manual", page=0, page_size=ADMIN_SERVICE_PAGE_SIZE),
    )


@router.message(F.text == "🕒 Свободные окна")
async def admin_availability_start_handler(message: types.Message, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return

    services = await database.get_all_services()
    if not services:
        await message.answer("⚙️ Сначала добавьте хотя бы одну услугу.", parse_mode="HTML")
        return

    await state.clear()
    await state.set_state(AdminAvailabilityForm.service_id)
    await state.update_data(availability_page=0)
    await message.answer(
        "🕒 <b>Свободные окна</b>\n\nВыберите услугу, чтобы посмотреть доступные слоты.",
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_service_picker_keyboard(services, prefix="avail", page=0, page_size=ADMIN_SERVICE_PAGE_SIZE),
    )


@router.message(Command("export_excel"))
@router.message(F.text == "📃 Excel")
async def export_excel_handler(message: types.Message):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return

    bookings = await database.get_all_bookings()
    if not bookings:
        await message.answer("📃 <b>Нет записей для выгрузки</b>", parse_mode="HTML")
        return

    file_path = "bookings_export.xlsx"
    _build_bookings_workbook(file_path, bookings)

    excel_file = FSInputFile(file_path)
    caption = (
        f"📃 <b>Экспорт записей</b>\n"
        f"<b>Салон:</b> {escape(salon_config.get('salon_name', 'Салон'))}\n"
        f"<b>Всего записей:</b> {len(bookings)}"
    )
    await message.answer_document(excel_file, caption=caption, parse_mode="HTML")
    os.remove(file_path)


@router.message(F.text == "🗓 Все записи")
async def view_all_handler(message: types.Message):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return
    bookings = await database.get_all_bookings_detailed()
    await _show_booking_list(message, context="all", page=_get_last_page(bookings))


@router.message(F.text == "🗓 На сегодня")
async def todays_bookings_handler(message: types.Message):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return
    await _show_booking_list(message, context="today", page=0)


@router.callback_query(F.data.startswith("manual_service_page_"))
async def manual_service_page_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    page = int(callback.data.rsplit("_", 1)[-1])
    services = await database.get_all_services()
    await state.set_state(ManualBookingForm.service_id)
    await state.update_data(service_page=page)
    await callback.message.edit_text(
        "➕ <b>Ручная запись</b>\n\nВыберите услугу для новой записи.",
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_service_picker_keyboard(services, prefix="manual", page=page, page_size=ADMIN_SERVICE_PAGE_SIZE),
    )


@router.callback_query(F.data.startswith("manual_service_"))
async def manual_service_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    _, _, page_str, service_id_str = callback.data.split("_", 3)
    services = await database.get_all_services()
    service = _find_service_by_id(services, int(service_id_str))
    if not service:
        await callback.answer("Услуга не найдена", show_alert=True)
        return

    options = await _build_admin_date_options(int(service.get("duration") or 60))
    await state.set_state(ManualBookingForm.date)
    await state.update_data(service_id=service["id"], service_page=int(page_str))
    if not options:
        await callback.message.edit_text(
            "➕ <b>Ручная запись</b>\n\nДля этой услуги сейчас нет доступных дат в окне записи.",
            parse_mode="HTML",
            reply_markup=keyboards.get_cancel_admin_action_keyboard(back_callback=f"manual_service_page_{page_str}", back_text="⬅️ Назад к услугам"),
        )
        return

    await callback.message.edit_text(
        (
            "➕ <b>Ручная запись</b>\n\n"
            f"<b>Услуга:</b> {escape(service['name'])}\n"
            "Выберите дату."
        ),
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_date_picker_keyboard(
            options,
            prefix="manual",
            back_callback=f"manual_service_page_{page_str}",
        ),
    )


@router.callback_query(F.data.startswith("manual_date_"))
async def manual_date_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    date_value = callback.data.replace("manual_date_", "", 1)
    data = await state.get_data()
    services = await database.get_all_services()
    service = _find_service_by_id(services, data.get("service_id"))
    if not service:
        await callback.answer("Услуга не найдена", show_alert=True)
        return

    times = await _build_admin_time_options(date_value, int(service.get("duration") or 60))
    await state.set_state(ManualBookingForm.time)
    await state.update_data(date=date_value)
    if not times:
        await callback.message.edit_text(
            "➕ <b>Ручная запись</b>\n\nНа эту дату подходящих слотов нет. Выберите другую дату.",
            parse_mode="HTML",
            reply_markup=keyboards.get_admin_date_picker_keyboard(
                await _build_admin_date_options(int(service.get("duration") or 60)),
                prefix="manual",
                back_callback=f"manual_service_page_{data.get('service_page', 0)}",
            ),
        )
        return

    await callback.message.edit_text(
        (
            "➕ <b>Ручная запись</b>\n\n"
            f"<b>Услуга:</b> {escape(service['name'])}\n"
            f"<b>Дата:</b> {escape(date_value)}\n"
            "Выберите время."
        ),
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_time_picker_keyboard(
            date_value,
            times,
            prefix="manual",
            back_callback="manual_date_back",
        ),
    )


@router.callback_query(F.data == "manual_date_back")
async def manual_date_back_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    data = await state.get_data()
    services = await database.get_all_services()
    service = _find_service_by_id(services, data.get("service_id"))
    if not service:
        await callback.answer("Услуга не найдена", show_alert=True)
        return

    options = await _build_admin_date_options(int(service.get("duration") or 60))
    await state.set_state(ManualBookingForm.date)
    await callback.message.edit_text(
        (
            "➕ <b>Ручная запись</b>\n\n"
            f"<b>Услуга:</b> {escape(service['name'])}\n"
            "Выберите дату."
        ),
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_date_picker_keyboard(
            options,
            prefix="manual",
            back_callback=f"manual_service_page_{data.get('service_page', 0)}",
        ),
    )


@router.callback_query(F.data.startswith("manual_time_"))
async def manual_time_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    _, _, date_value, time_value = callback.data.split("_", 3)
    await state.set_state(ManualBookingForm.phone)
    await state.update_data(date=date_value, time=time_value)
    await callback.message.answer(
        (
            "➕ <b>Ручная запись</b>\n\n"
            f"<b>Дата:</b> {escape(date_value)}\n"
            f"<b>Время:</b> {escape(time_value)}\n\n"
            "Введите телефон клиента в формате +7 или нажмите «Пропустить»."
        ),
        parse_mode="HTML",
        reply_markup=_build_manual_phone_keyboard(),
    )

@router.message(ManualBookingForm.name)
async def manual_booking_name_handler(message: types.Message, state):
    name = (message.text or "").strip()
    if len(name) < 2:
        await message.answer("Введите имя клиента не короче 2 символов.")
        return

    await state.set_state(ManualBookingForm.source)
    await state.update_data(name=name)
    await message.answer(
        "Выберите источник записи.",
        reply_markup=keyboards.get_manual_booking_source_keyboard(),
    )

@router.message(ManualBookingForm.phone)
async def manual_booking_phone_handler(message: types.Message, state):
    phone = normalize_phone(message.text)
    if not phone:
        await message.answer("Телефон не похож на корректный номер. Используйте формат +7 или нажмите «Пропустить».")
        return

    snapshot = await database.get_client_snapshot_by_phone(phone)
    await state.set_state(ManualBookingForm.name)
    await state.update_data(phone=phone)
    if snapshot:
        await state.update_data(found_name=snapshot["name"])
        await message.answer(
            (
                "👤 <b>Клиент найден</b>\n\n"
                f"<b>Имя:</b> {escape(snapshot['name'])}\n"
                f"<b>Последняя запись:</b> {escape(snapshot['last_date'] or '—')}\n"
                f"<b>Записей всего:</b> {int(snapshot['total_bookings'])}\n\n"
                "Можно использовать найденное имя или ввести новое."
            ),
            parse_mode="HTML",
            reply_markup=_build_manual_name_keyboard(snapshot["name"]),
        )
        return

    await message.answer("Введите имя клиента.")


@router.callback_query(F.data == "manual_phone_skip")
async def manual_phone_skip_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    await state.set_state(ManualBookingForm.name)
    await state.update_data(phone="", found_name=None)
    await callback.message.answer("Введите имя клиента.")


@router.callback_query(F.data == "manual_use_found_name")
async def manual_use_found_name_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    data = await state.get_data()
    found_name = (data.get("found_name") or "").strip()
    if len(found_name) < 2:
        await callback.answer("Не удалось использовать найденное имя.", show_alert=True)
        return

    await state.set_state(ManualBookingForm.source)
    await state.update_data(name=found_name)
    await callback.message.answer(
        "Выберите источник записи.",
        reply_markup=keyboards.get_manual_booking_source_keyboard(),
    )


@router.callback_query(F.data == "manual_enter_other_name")
async def manual_enter_other_name_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    await state.set_state(ManualBookingForm.name)
    await callback.message.answer("Введите имя клиента.")

@router.callback_query(F.data.startswith("manual_source_"))
async def manual_booking_source_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    source = callback.data.replace("manual_source_", "", 1)
    await state.set_state(ManualBookingForm.notes)
    await state.update_data(source=source)
    await callback.message.answer(
        "Добавьте комментарий к записи или нажмите «Пропустить».",
        reply_markup=keyboards.get_manual_booking_notes_keyboard(),
    )


@router.callback_query(F.data == "manual_notes_skip")
async def manual_booking_notes_skip_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    await state.update_data(notes="")
    await state.set_state(ManualBookingForm.confirm)
    await _show_manual_booking_confirmation(callback.message, state)


async def _show_manual_booking_confirmation(message: types.Message, state):
    data = await state.get_data()
    services = await database.get_all_services()
    service = _find_service_by_id(services, data.get("service_id"))
    if not service:
        await message.answer("Не удалось собрать карточку записи: услуга не найдена.")
        return

    text = (
        "📝 <b>Подтверждение ручной записи</b>\n\n"
        f"<b>Услуга:</b> {escape(service['name'])}\n"
        f"<b>Дата:</b> {escape(data['date'])}\n"
        f"<b>Время:</b> {escape(data['time'])}\n"
        f"<b>Клиент:</b> {escape(data['name'])}\n"
        f"<b>Телефон:</b> {escape(data.get('phone') or '—')}\n"
        f"<b>Источник:</b> {escape(_source_label(data.get('source')))}\n"
        f"<b>Комментарий:</b> {escape(data.get('notes') or '—')}\n"
        f"<b>Стоимость:</b> {escape(format_money(service.get('price_value') or 0))}"
    )
    await message.answer(text, parse_mode="HTML", reply_markup=keyboards.get_manual_booking_confirm_keyboard())


@router.message(ManualBookingForm.notes)
async def manual_booking_notes_handler(message: types.Message, state):
    notes = (message.text or "").strip()
    await state.update_data(notes=notes)
    await state.set_state(ManualBookingForm.confirm)
    await _show_manual_booking_confirmation(message, state)


@router.callback_query(F.data == "manual_confirm_submit")
async def manual_booking_confirm_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    data = await state.get_data()
    services = await database.get_all_services()
    service = _find_service_by_id(services, data.get("service_id"))
    if not service:
        await callback.message.answer("Услуга не найдена. Начните создание записи заново.")
        await state.clear()
        return

    linked_user_id = None
    if data.get("phone"):
        linked_user_id = await database.get_existing_user_id_by_phone(data["phone"])

    created = await database.create_manual_booking(
        name=data["name"],
        phone=(data.get("phone") or None),
        date=data["date"],
        time=data["time"],
        duration=int(service.get("duration") or 60),
        service_name=service["name"],
        price=int(service.get("price_value") or 0),
        source=str(data.get("source") or "manual"),
        notes=(data.get("notes") or "").strip() or None,
    )
    if not created:
        times = await _build_admin_time_options(data["date"], int(service.get("duration") or 60))
        await state.set_state(ManualBookingForm.time)
        await callback.message.answer(
            "Этот слот уже заняли. Выберите другое время.",
            reply_markup=keyboards.get_admin_time_picker_keyboard(
                data["date"],
                times,
                prefix="manual",
                back_callback="manual_date_back",
            ),
        )
        return

    await _notify_client_about_admin_booking_action(
        callback.bot,
        user_id=linked_user_id,
        title="📅 <b>Запись оформлена</b>",
        lines=[
            f"<b>Услуга:</b> {escape(service['name'])}",
            f"<b>Дата:</b> {escape(data['date'])}",
            f"<b>Время:</b> {escape(data['time'])}",
            f"<b>Телефон:</b> {escape(data.get('phone') or '—')}",
            "Ждём вас в выбранное время.",
        ],
    )

    await state.clear()
    await callback.message.answer(
        (
            "✅ <b>Запись добавлена</b>\n\n"
            f"<b>Клиент:</b> {escape(data['name'])}\n"
            f"<b>Услуга:</b> {escape(service['name'])}\n"
            f"<b>Когда:</b> {escape(data['date'])} в {escape(data['time'])}\n"
            f"<b>Источник:</b> {escape(_source_label(data.get('source')))}"
        ),
        parse_mode="HTML",
    )


@router.callback_query(F.data.startswith("avail_service_page_"))
async def availability_service_page_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    page = int(callback.data.rsplit("_", 1)[-1])
    services = await database.get_all_services()
    await state.set_state(AdminAvailabilityForm.service_id)
    await state.update_data(availability_page=page)
    await callback.message.edit_text(
        "🕒 <b>Свободные окна</b>\n\nВыберите услугу, чтобы посмотреть доступные слоты.",
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_service_picker_keyboard(services, prefix="avail", page=page, page_size=ADMIN_SERVICE_PAGE_SIZE),
    )


@router.callback_query(F.data.startswith("avail_service_"))
async def availability_service_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    _, _, page_str, service_id_str = callback.data.split("_", 3)
    services = await database.get_all_services()
    service = _find_service_by_id(services, int(service_id_str))
    if not service:
        await callback.answer("Услуга не найдена", show_alert=True)
        return

    options = await _build_admin_date_options(int(service.get("duration") or 60))
    await state.set_state(AdminAvailabilityForm.date)
    await state.update_data(service_id=service["id"], availability_page=int(page_str))
    await callback.message.edit_text(
        (
            "🕒 <b>Свободные окна</b>\n\n"
            f"<b>Услуга:</b> {escape(service['name'])}\n"
            "Выберите дату."
        ),
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_date_picker_keyboard(
            options,
            prefix="avail",
            back_callback=f"avail_service_page_{page_str}",
        ),
    )


@router.callback_query(F.data.startswith("avail_date_"))
async def availability_date_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    date_value = callback.data.replace("avail_date_", "", 1)
    data = await state.get_data()
    services = await database.get_all_services()
    service = _find_service_by_id(services, data.get("service_id"))
    if not service:
        await callback.answer("Услуга не найдена", show_alert=True)
        return

    times = await _build_admin_time_options(date_value, int(service.get("duration") or 60))
    if not times:
        await callback.message.answer("На эту дату свободных слотов нет.")
        return

    await callback.message.edit_text(
        (
            "🕒 <b>Свободные окна</b>\n\n"
            f"<b>Услуга:</b> {escape(service['name'])}\n"
            f"<b>Дата:</b> {escape(date_value)}\n\n"
            f"{escape(', '.join(times))}"
        ),
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_date_picker_keyboard(
            await _build_admin_date_options(int(service.get("duration") or 60)),
            prefix="avail",
            back_callback=f"avail_service_page_{data.get('availability_page', 0)}",
        ),
    )


@router.message(F.text == "🔎 Поиск")
async def search_bookings_handler(message: types.Message, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return
    await state.set_state(SearchBookingForm.query)
    await message.answer(
        "🔎 <b>Поиск записи</b>\n\nВведите имя, телефон, дату, время, услугу или источник записи. Например: <code>777</code>, <code>WhatsApp</code> или <code>21.03.2026</code>.",
        parse_mode="HTML",
    )


@router.message(SearchBookingForm.query)
async def process_booking_search(message: types.Message, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return

    query = (message.text or "").strip()
    if len(query) < 2:
        await message.answer("Введите хотя бы 2 символа для поиска.")
        return

    results = await database.search_bookings(query, limit=10)
    await state.clear()
    if not results:
        await message.answer("🔎 <b>Ничего не найдено</b>\n\nПопробуйте другой запрос.", parse_mode="HTML")
        return

    lines = [f"🔎 <b>Результаты поиска</b>", f"<i>Запрос: {escape(query)}</i>", ""]
    for booking_id, name, phone, date, time, status in results:
        lines.append(
            f"<b>#{booking_id}</b> {escape(name)}\n"
            f"Статус: {escape(_status_label(status))}\n"
            f"Когда: {escape(date)} в {escape(time)}\n"
            f"Телефон: {escape(phone)}\n"
        )
    await message.answer("\n".join(lines), parse_mode="HTML")


@router.callback_query(F.data.startswith("bookings_page_"))
async def bookings_page_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    _, _, context, page_str = callback.data.split("_", 3)
    await _show_booking_list(callback, context=context, page=int(page_str))


@router.callback_query(F.data.startswith("booking_actions_"))
async def booking_actions_callback_v2(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    _, _, context, page_str, booking_id_str = callback.data.split("_", 4)
    booking = await database.get_booking_admin_details(int(booking_id_str))
    if not booking:
        await callback.answer("Запись не найдена", show_alert=True)
        return

    if len(booking) == 13:
        booking_id, user_id, name, phone, date, time, status, duration, service_name, price, source, notes, created_by_admin = booking
        created_at = None
    else:
        booking_id, user_id, name, phone, date, time, status, duration, service_name, price, source, notes, created_by_admin, created_at = booking
    text = (
        "📝 <b>Карточка записи</b>\n\n"
        f"<b>Клиент:</b> {escape(name)}\n"
        f"<b>Телефон:</b> {escape(phone)}\n"
        f"<b>Услуга:</b> {escape(service_name or '—')}\n"
        f"<b>Дата:</b> {escape(date)}\n"
        f"<b>Время:</b> {escape(time)}\n"
        f"<b>Длительность:</b> {int(duration or 0)} мин\n"
        f"<b>Сумма:</b> {escape(format_money(price))}\n"
        f"<b>Источник:</b> {escape(_source_label(source))}\n"
        f"<b>Создана админом:</b> {'Да' if int(created_by_admin or 0) else 'Нет'}\n"
        f"<b>Создана:</b> {escape(_format_iso_to_date(created_at))}\n"
        f"<b>Статус:</b> {escape(_status_label(status))}\n"
        f"<b>Комментарий:</b> {escape(notes or '—')}"
    )
    await callback.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_booking_actions_keyboard(
            booking_id,
            phone,
            context,
            int(page_str),
            status=status,
            telegram_user_id=user_id,
        ),
    )


@router.callback_query(F.data.startswith("booking_actions_legacy_"))
async def booking_actions_callback_legacy(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    _, _, context, page_str, booking_id_str = callback.data.split("_", 4)
    booking = await database.get_booking_record_by_id(int(booking_id_str))
    if not booking:
        await callback.answer("Запись не найдена", show_alert=True)
        return
    booking_id, user_id, name, phone, date, time, status, _duration = booking
    text = (
        "📝 <b>Карточка записи</b>\n\n"
        f"<b>Клиент:</b> {escape(name)}\n"
        f"<b>Телефон:</b> {escape(phone)}\n"
        f"<b>Дата:</b> {escape(date)}\n"
        f"<b>Время:</b> {escape(time)}\n"
        f"<b>Статус:</b> {escape(_status_label(status))}"
    )
    await callback.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=keyboards.get_admin_booking_actions_keyboard(
            booking_id,
            phone,
            context,
            int(page_str),
            status=status,
            telegram_user_id=user_id,
        ),
    )


@router.callback_query(F.data.startswith("show_phone_"))
async def show_phone_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    phone = callback.data.replace("show_phone_", "", 1)
    await callback.answer(f"Номер: +{phone}", show_alert=True)


@router.callback_query(F.data.startswith("admin_booking_status_"))
async def admin_booking_status_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    prefix = "admin_booking_status_"
    payload = callback.data[len(prefix):]
    first_separator = payload.find("_")
    booking_id_str = payload[:first_separator]
    status, context, page_str = payload[first_separator + 1 :].rsplit("_", 2)

    booking = await database.get_booking_admin_details(int(booking_id_str)) if status == "cancelled" else None
    await database.update_booking_status(int(booking_id_str), status)
    if status == "cancelled" and booking:
        if len(booking) == 13:
            _booking_id, user_id, name, phone, date, time, _current_status, _duration, service_name, _price, _source, _notes, _created_by_admin = booking
        else:
            _booking_id, user_id, name, phone, date, time, _current_status, _duration, service_name, _price, _source, _notes, _created_by_admin, _created_at = booking
        await _notify_client_about_admin_booking_action(
            callback.bot,
            user_id=user_id,
            title="❌ <b>Ваша запись отменена</b>",
            lines=[
                f"<b>Услуга:</b> {escape(service_name or '—')}",
                f"<b>Дата:</b> {escape(date)}",
                f"<b>Время:</b> {escape(time)}",
                "Если захотите, можно записаться снова.",
            ],
        )
    await callback.answer("Статус обновлен")
    await _show_booking_list(callback, context=context, page=int(page_str))
booking_actions_callback = booking_actions_callback_v2


def _short_name(name: str, limit: int = 18) -> str:
    value = (name or "").strip()
    if len(value) <= limit:
        return value
    return value[: limit - 1].rstrip() + "..."


def _normalize_context_value(value: str | None) -> str | None:
    if not value or value == "-":
        return None
    return value


def _build_booking_filters_markup(bookings, context: str, page: int, total_pages: int, *, source_filter: str = "all", date_value: str | None = None):
    from aiogram.utils.keyboard import InlineKeyboardBuilder

    builder = InlineKeyboardBuilder()
    status_prefix = {
        "scheduled": "🟢",
        "completed": "✅",
        "no_show": "🟠",
        "cancelled": "❌",
    }
    date_token = date_value or "-"

    for booking_id, name, _phone, date, time, _price, status in bookings:
        label = f"{status_prefix.get(status, '•')} {time} · {_short_name(name)}"
        if context == "all":
            label = f"{status_prefix.get(status, '•')} {date} · {time} · {_short_name(name)}"
        builder.row(
            types.InlineKeyboardButton(
                text=label,
                callback_data=f"booking_actions|{context}|{date_token}|{source_filter}|{page}|{booking_id}",
            )
        )

    filter_options = [
        ("Все", "all"),
        ("TG", "telegram"),
        ("WA", "whatsapp"),
        ("IG", "instagram"),
        ("📞", "phone"),
        ("🏠", "offline"),
        ("✍️", "manual"),
    ]
    first_row = []
    second_row = []
    for index, (label, value) in enumerate(filter_options):
        display = f"· {label}" if value == source_filter else label
        button = types.InlineKeyboardButton(
            text=display,
            callback_data=f"bookings_filter|{context}|{date_token}|{value}",
        )
        if index < 4:
            first_row.append(button)
        else:
            second_row.append(button)
    builder.row(*first_row)
    builder.row(*second_row)

    nav_buttons = []
    if page > 0:
        nav_buttons.append(
            types.InlineKeyboardButton(
                text="⬅️ Назад",
                callback_data=f"bookings_page|{context}|{date_token}|{source_filter}|{page - 1}",
            )
        )
    if page < total_pages - 1:
        nav_buttons.append(
            types.InlineKeyboardButton(
                text="Вперёд ➡️",
                callback_data=f"bookings_page|{context}|{date_token}|{source_filter}|{page + 1}",
            )
        )
    if nav_buttons:
        builder.row(*nav_buttons)

    builder.row(types.InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_admin_action"))
    return builder.as_markup()


def _build_booking_actions_markup(
    booking_id: int,
    phone: str,
    *,
    context: str,
    page: int,
    status: str,
    telegram_user_id: int | None,
    source_filter: str = "all",
    date_value: str | None = None,
):
    from aiogram.utils.keyboard import InlineKeyboardBuilder

    builder = InlineKeyboardBuilder()
    digits = normalize_phone(phone) or ""
    digits = digits.replace("+", "")
    date_token = date_value or "-"

    contact_buttons = []
    if digits:
        contact_buttons.append(types.InlineKeyboardButton(text="💬 Написать", url=f"https://wa.me/{digits}"))
    if digits:
        contact_buttons.append(types.InlineKeyboardButton(text="📞 Показать номер", callback_data=f"show_phone_{digits}"))
    if contact_buttons:
        builder.row(*contact_buttons)

    builder.row(
        types.InlineKeyboardButton(text="✏️ Имя", callback_data=f"booking_edit_name|{booking_id}|{context}|{date_token}|{source_filter}|{page}"),
        types.InlineKeyboardButton(text="📱 Телефон", callback_data=f"booking_edit_phone|{booking_id}|{context}|{date_token}|{source_filter}|{page}"),
    )
    builder.row(
        types.InlineKeyboardButton(text="📍 Источник", callback_data=f"booking_edit_source|{booking_id}|{context}|{date_token}|{source_filter}|{page}"),
        types.InlineKeyboardButton(text="📝 Комментарий", callback_data=f"booking_edit_notes|{booking_id}|{context}|{date_token}|{source_filter}|{page}"),
    )
    if status == "scheduled":
        builder.row(
            types.InlineKeyboardButton(
                text="🔁 Перенести",
                callback_data=f"booking_reschedule|{booking_id}|{context}|{date_token}|{source_filter}|{page}",
            )
        )

    if status == "scheduled":
        builder.row(
            types.InlineKeyboardButton(
                text="✅ Отметить выполненной",
                callback_data=f"adminstatus|{booking_id}|completed|{context}|{date_token}|{source_filter}|{page}",
            ),
            types.InlineKeyboardButton(
                text="🟠 Не пришел",
                callback_data=f"adminstatus|{booking_id}|no_show|{context}|{date_token}|{source_filter}|{page}",
            ),
        )
        builder.row(
            types.InlineKeyboardButton(
                text="❌ Отменить запись",
                callback_data=f"adminstatus|{booking_id}|cancelled|{context}|{date_token}|{source_filter}|{page}",
            )
        )
    else:
        builder.row(
            types.InlineKeyboardButton(
                text="🔄 Вернуть в активные",
                callback_data=f"adminstatus|{booking_id}|scheduled|{context}|{date_token}|{source_filter}|{page}",
            )
        )

    builder.row(
        types.InlineKeyboardButton(
            text="⬅️ Назад к списку",
            callback_data=f"bookings_page|{context}|{date_token}|{source_filter}|{page}",
        )
    )
    return builder.as_markup()


def _build_edit_source_markup(booking_id: int, *, context: str, page: int, source_filter: str = "all", date_value: str | None = None):
    from aiogram.utils.keyboard import InlineKeyboardBuilder

    builder = InlineKeyboardBuilder()
    date_token = date_value or "-"
    for label, value in (
        ("WhatsApp", "whatsapp"),
        ("Instagram", "instagram"),
        ("Звонок", "phone"),
        ("Telegram", "telegram"),
        ("Офлайн", "offline"),
        ("Вручную", "manual"),
    ):
        builder.row(
            types.InlineKeyboardButton(
                text=label,
                callback_data=f"booking_set_source|{booking_id}|{value}|{context}|{date_token}|{source_filter}|{page}",
            )
        )
    builder.row(
        types.InlineKeyboardButton(
            text="⬅️ Назад",
            callback_data=f"booking_actions|{context}|{date_token}|{source_filter}|{page}|{booking_id}",
        )
    )
    builder.row(types.InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_admin_action"))
    return builder.as_markup()


def _build_manual_phone_keyboard():
    from aiogram.utils.keyboard import InlineKeyboardBuilder

    builder = InlineKeyboardBuilder()
    builder.row(types.InlineKeyboardButton(text="Пропустить", callback_data="manual_phone_skip"))
    builder.row(types.InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_admin_action"))
    return builder.as_markup()


def _build_manual_name_keyboard(found_name: str):
    from aiogram.utils.keyboard import InlineKeyboardBuilder

    builder = InlineKeyboardBuilder()
    builder.row(types.InlineKeyboardButton(text=f"Использовать: {found_name}", callback_data="manual_use_found_name"))
    builder.row(types.InlineKeyboardButton(text="Ввести другое имя", callback_data="manual_enter_other_name"))
    builder.row(types.InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_admin_action"))
    return builder.as_markup()


async def _show_booking_card(message_or_callback, booking_id: int, *, context: str, page: int, source_filter: str = "all", date_value: str | None = None):
    booking = await database.get_booking_admin_details(int(booking_id))
    if not booking:
        if isinstance(message_or_callback, types.CallbackQuery):
            await message_or_callback.answer("Запись не найдена", show_alert=True)
        else:
            await message_or_callback.answer("Запись не найдена.")
        return

    if len(booking) == 13:
        booking_id, user_id, name, phone, date, time, status, duration, service_name, price, source, notes, created_by_admin = booking
        created_at = None
    else:
        booking_id, user_id, name, phone, date, time, status, duration, service_name, price, source, notes, created_by_admin, created_at = booking
    text = (
        "📝 <b>Карточка записи</b>\n\n"
        f"<b>Клиент:</b> {escape(name)}\n"
        f"<b>Телефон:</b> {escape(phone)}\n"
        f"<b>Услуга:</b> {escape(service_name or '—')}\n"
        f"<b>Дата:</b> {escape(date)}\n"
        f"<b>Время:</b> {escape(time)}\n"
        f"<b>Длительность:</b> {int(duration or 0)} мин\n"
        f"<b>Сумма:</b> {escape(format_money(price))}\n"
        f"<b>Источник:</b> {escape(_source_label(source))}\n"
        f"<b>Создана админом:</b> {'Да' if int(created_by_admin or 0) else 'Нет'}\n"
        f"<b>Создана:</b> {escape(_format_iso_to_date(created_at))}\n"
        f"<b>Статус:</b> {escape(_status_label(status))}\n"
        f"<b>Комментарий:</b> {escape(notes or '—')}"
    )
    reply_markup = _build_booking_actions_markup(
        booking_id,
        phone,
        context=context,
        page=page,
        status=status,
        telegram_user_id=user_id,
        source_filter=source_filter,
        date_value=date_value,
    )
    if isinstance(message_or_callback, types.CallbackQuery):
        await message_or_callback.message.edit_text(text, parse_mode="HTML", reply_markup=reply_markup)
    else:
        await message_or_callback.answer(text, parse_mode="HTML", reply_markup=reply_markup)


def _render_booking_page(bookings, title: str, page: int, *, source_filter: str = "all"):
    page_items, page, total_pages = _paginate(bookings, page)
    lines = [f"🗓 <b>{title}</b>", f"<i>Страница {page + 1} из {total_pages}</i>", ""]
    if source_filter != "all":
        lines.extend([f"<b>Фильтр:</b> {_source_label(source_filter)}", ""])

    for idx, (_booking_id, name, phone, date, time, price, status) in enumerate(page_items, start=1 + page * BOOKINGS_PAGE_SIZE):
        safe_name = escape(name)
        safe_phone = escape(phone)
        safe_date = escape(date)
        safe_time = escape(time)
        status_badge = {
            "scheduled": "🟢 Активна",
            "completed": "✅ Выполнена",
            "no_show": "🟠 Не пришел",
            "cancelled": "❌ Отменена",
        }.get(status, escape(_status_label(status)))
        lines.append(
            f"┌ <b>Запись #{idx}</b>\n"
            f"├ <b>Когда:</b> {safe_date} в {safe_time}\n"
            f"├ <b>Клиент:</b> {safe_name}\n"
            f"├ <b>Телефон:</b> {safe_phone}\n"
            f"├ <b>Статус:</b> {status_badge}\n"
            f"└ <b>Сумма:</b> {escape(format_money(price))}\n"
        )
    if not page_items:
        lines.append("Записей пока нет.")
    return "\n".join(lines), page_items, page, total_pages


def _render_clients_page(clients: list[tuple], page: int):
    page_items, page, total_pages = _paginate(clients, page, page_size=CLIENTS_PAGE_SIZE)
    lines = [f"👥 <b>Клиенты</b>", f"<i>Страница {page + 1} из {total_pages}</i>", ""]

    for idx, (_phone_key, name, phone, total, completed, cancelled, no_show, revenue, last_date_iso, _last_created_at) in enumerate(
        page_items,
        start=1 + page * CLIENTS_PAGE_SIZE,
    ):
        last_visit = "—"
        if last_date_iso:
            try:
                last_visit = datetime.fromisoformat(last_date_iso).strftime("%d.%m.%Y")
            except ValueError:
                last_visit = last_date_iso

        lines.append(
            f"{idx}. <b>{escape(name or 'Без имени')}</b>\n"
            f"Телефон: {escape(phone or '—')}\n"
            f"Записей: <b>{int(total or 0)}</b> · Выполнено: <b>{int(completed or 0)}</b>\n"
            f"Сумма: <b>{escape(format_money(revenue))}</b> · Последняя запись: <b>{escape(last_visit)}</b>\n"
        )

    if not page_items:
        lines.append("Клиентская база пока пуста.")

    return "\n".join(lines), page_items, page, total_pages


def _build_clients_keyboard(page_items: list[tuple], page: int, total_pages: int):
    from aiogram.utils.keyboard import InlineKeyboardBuilder

    builder = InlineKeyboardBuilder()
    for index, (_phone_key, name, phone, total, _completed, _cancelled, _no_show, _revenue, _last_date_iso, _last_created_at) in enumerate(page_items):
        label = f"{name or 'Без имени'}"
        if phone:
            label += f" · {phone}"
        else:
            label += f" · {int(total or 0)} записей"
        builder.row(types.InlineKeyboardButton(text=label[:64], callback_data=f"clients_open|{page}|{index}"))

    nav_row = []
    if page > 0:
        nav_row.append(types.InlineKeyboardButton(text="⬅️", callback_data=f"clients_page|{page - 1}"))
    nav_row.append(types.InlineKeyboardButton(text=f"{page + 1}/{total_pages}", callback_data="clients_page|noop"))
    if page < total_pages - 1:
        nav_row.append(types.InlineKeyboardButton(text="➡️", callback_data=f"clients_page|{page + 1}"))
    builder.row(*nav_row)
    return builder.as_markup()


def _format_client_card(client: tuple) -> str:
    _phone_key, name, phone, total, completed, cancelled, no_show, revenue, last_date_iso, last_created_at = client
    last_visit = "—"
    if last_date_iso:
        try:
            last_visit = datetime.fromisoformat(last_date_iso).strftime("%d.%m.%Y")
        except ValueError:
            last_visit = last_date_iso

    return (
        "👤 <b>Карточка клиента</b>\n\n"
        f"<b>Имя:</b> {escape(name or 'Без имени')}\n"
        f"<b>Телефон:</b> {escape(phone or '—')}\n"
        f"<b>Всего записей:</b> {int(total or 0)}\n"
        f"<b>Выполнено:</b> {int(completed or 0)}\n"
        f"<b>Отменено:</b> {int(cancelled or 0)}\n"
        f"<b>Не пришел:</b> {int(no_show or 0)}\n"
        f"<b>Сумма по выполненным:</b> {escape(format_money(revenue))}\n"
        f"<b>Последняя запись:</b> {escape(last_visit)}\n"
        f"<b>Последнее создание:</b> {escape(_format_iso_to_date(last_created_at))}"
    )


async def _show_booking_list(message_or_callback, *, context: str, page: int = 0, source_filter: str = "all", target_date: str | None = None):
    await database.sync_completed_bookings()
    if context == "today":
        today_str = datetime.now().strftime("%d.%m.%Y")
        bookings = await database.get_bookings_by_date_detailed_filtered(today_str, source_filter)
        title = f"Записи на сегодня ({today_str})"
        date_value = today_str
    elif context == "date" and target_date:
        bookings = await database.get_bookings_by_date_detailed_filtered(target_date, source_filter)
        title = f"Записи на дату ({target_date})"
        date_value = target_date
    else:
        bookings = await database.get_all_bookings_detailed_filtered(source_filter)
        title = "Все записи"
        date_value = None

    text, page_items, page, total_pages = _render_booking_page(bookings, title, page, source_filter=source_filter)
    markup = _build_booking_filters_markup(
        page_items,
        context,
        page,
        total_pages,
        source_filter=source_filter,
        date_value=date_value,
    )

    if isinstance(message_or_callback, types.CallbackQuery):
        await message_or_callback.message.edit_text(text, parse_mode="HTML", reply_markup=markup)
    else:
        await message_or_callback.answer(text, parse_mode="HTML", reply_markup=markup)


@router.message(F.text == "📅 По дате")
async def bookings_by_date_handler(message: types.Message, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return
    await state.set_state(AdminBookingsByDateForm.date)
    await message.answer("📅 <b>Просмотр по дате</b>\n\nВведите дату в формате <code>30.03.2026</code>.", parse_mode="HTML")


@router.message(AdminBookingsByDateForm.date)
async def bookings_by_date_value_handler(message: types.Message, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(message.from_user.id) != admin_id:
        return
    target_date = (message.text or "").strip()
    if target_date in MENU_ESCAPE_TEXTS:
        await state.clear()
        await message.answer("Сценарий просмотра по дате сброшен. Нажмите кнопку ещё раз.")
        return
    if not _safe_parse_date(target_date):
        await message.answer("Введите дату в формате <code>дд.мм.гггг</code>.", parse_mode="HTML")
        return
    await state.clear()
    await _show_booking_list(message, context="date", page=0, source_filter="all", target_date=target_date)


@router.callback_query(F.data.startswith("bookings_page|"))
async def bookings_page_callback_v2(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    _, context, date_token, source_filter, page_str = callback.data.split("|", 4)
    await _show_booking_list(
        callback,
        context=context,
        page=int(page_str),
        source_filter=source_filter,
        target_date=_normalize_context_value(date_token),
    )


@router.callback_query(F.data.startswith("bookings_filter|"))
async def bookings_filter_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    _, context, date_token, source_filter = callback.data.split("|", 3)
    await _show_booking_list(
        callback,
        context=context,
        page=0,
        source_filter=source_filter,
        target_date=_normalize_context_value(date_token),
    )


@router.callback_query(F.data.startswith("booking_actions|"))
async def booking_actions_callback_pipe(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    _, context, date_token, source_filter, page_str, booking_id_str = callback.data.split("|", 5)
    await _show_booking_card(
        callback,
        int(booking_id_str),
        context=context,
        page=int(page_str),
        source_filter=source_filter,
        date_value=_normalize_context_value(date_token),
    )


@router.callback_query(F.data.startswith("adminstatus|"))
async def admin_booking_status_callback_v2(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer("Статус обновлен")
    _, booking_id_str, status, context, date_token, source_filter, page_str = callback.data.split("|", 6)
    booking = await database.get_booking_admin_details(int(booking_id_str)) if status == "cancelled" else None
    await database.update_booking_status(int(booking_id_str), status)
    if status == "cancelled" and booking:
        if len(booking) == 13:
            _booking_id, user_id, name, phone, date, time, _current_status, _duration, service_name, _price, _source, _notes, _created_by_admin = booking
        else:
            _booking_id, user_id, name, phone, date, time, _current_status, _duration, service_name, _price, _source, _notes, _created_by_admin, _created_at = booking
        await _notify_client_about_admin_booking_action(
            callback.bot,
            user_id=user_id,
            title="❌ <b>Ваша запись отменена</b>",
            lines=[
                f"<b>Услуга:</b> {escape(service_name or '—')}",
                f"<b>Дата:</b> {escape(date)}",
                f"<b>Время:</b> {escape(time)}",
                "Если захотите, можно записаться снова.",
            ],
        )
    await _show_booking_list(
        callback,
        context=context,
        page=int(page_str),
        source_filter=source_filter,
        target_date=_normalize_context_value(date_token),
    )


@router.callback_query(F.data.startswith("booking_edit_name|"))
async def booking_edit_name_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    _, booking_id_str, context, date_token, source_filter, page_str = callback.data.split("|", 5)
    await state.set_state(AdminEditBookingForm.name)
    await state.update_data(
        edit_booking_id=int(booking_id_str),
        edit_context=context,
        edit_date_value=_normalize_context_value(date_token),
        edit_source_filter=source_filter,
        edit_page=int(page_str),
    )
    await callback.message.answer("✏️ Введите новое имя клиента.")


@router.callback_query(F.data.startswith("booking_edit_phone|"))
async def booking_edit_phone_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    _, booking_id_str, context, date_token, source_filter, page_str = callback.data.split("|", 5)
    await state.set_state(AdminEditBookingForm.phone)
    await state.update_data(
        edit_booking_id=int(booking_id_str),
        edit_context=context,
        edit_date_value=_normalize_context_value(date_token),
        edit_source_filter=source_filter,
        edit_page=int(page_str),
    )
    await callback.message.answer("📱 Введите новый телефон клиента в формате +7.")


@router.callback_query(F.data.startswith("booking_edit_notes|"))
async def booking_edit_notes_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    _, booking_id_str, context, date_token, source_filter, page_str = callback.data.split("|", 5)
    await state.set_state(AdminEditBookingForm.notes)
    await state.update_data(
        edit_booking_id=int(booking_id_str),
        edit_context=context,
        edit_date_value=_normalize_context_value(date_token),
        edit_source_filter=source_filter,
        edit_page=int(page_str),
    )
    await callback.message.answer("📝 Введите комментарий к записи. Отправьте <code>-</code>, чтобы очистить комментарий.", parse_mode="HTML")


@router.callback_query(F.data.startswith("booking_edit_source|"))
async def booking_edit_source_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer()
    _, booking_id_str, context, date_token, source_filter, page_str = callback.data.split("|", 5)
    await callback.message.edit_reply_markup(
        reply_markup=_build_edit_source_markup(
            int(booking_id_str),
            context=context,
            page=int(page_str),
            source_filter=source_filter,
            date_value=_normalize_context_value(date_token),
        )
    )


@router.callback_query(F.data.startswith("booking_set_source|"))
async def booking_set_source_callback(callback: types.CallbackQuery):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return
    await callback.answer("Источник обновлен")
    _, booking_id_str, source, context, date_token, source_filter, page_str = callback.data.split("|", 6)
    await database.update_booking_source(int(booking_id_str), source)
    await _show_booking_card(
        callback,
        int(booking_id_str),
        context=context,
        page=int(page_str),
        source_filter=source_filter,
        date_value=_normalize_context_value(date_token),
    )


@router.message(AdminEditBookingForm.name)
async def booking_edit_name_value_handler(message: types.Message, state):
    name = (message.text or "").strip()
    if len(name) < 2:
        await message.answer("Введите имя не короче 2 символов.")
        return
    data = await state.get_data()
    await database.update_booking_name(int(data["edit_booking_id"]), name)
    await state.clear()
    await message.answer("✅ Имя обновлено.")
    await _show_booking_card(
        message,
        int(data["edit_booking_id"]),
        context=data["edit_context"],
        page=int(data["edit_page"]),
        source_filter=data["edit_source_filter"],
        date_value=data.get("edit_date_value"),
    )


@router.message(AdminEditBookingForm.phone)
async def booking_edit_phone_value_handler(message: types.Message, state):
    phone = normalize_phone(message.text)
    if not phone:
        await message.answer("Телефон не похож на корректный номер. Используйте формат +7.")
        return
    data = await state.get_data()
    await database.update_booking_phone(int(data["edit_booking_id"]), phone)
    await state.clear()
    await message.answer("✅ Телефон обновлен.")
    await _show_booking_card(
        message,
        int(data["edit_booking_id"]),
        context=data["edit_context"],
        page=int(data["edit_page"]),
        source_filter=data["edit_source_filter"],
        date_value=data.get("edit_date_value"),
    )


@router.message(AdminEditBookingForm.notes)
async def booking_edit_notes_value_handler(message: types.Message, state):
    raw_notes = (message.text or "").strip()
    notes = None if raw_notes in {"", "-"} else raw_notes
    data = await state.get_data()
    await database.update_booking_notes(int(data["edit_booking_id"]), notes)
    await state.clear()
    await message.answer("✅ Комментарий обновлен.")
    await _show_booking_card(
        message,
        int(data["edit_booking_id"]),
        context=data["edit_context"],
        page=int(data["edit_page"]),
        source_filter=data["edit_source_filter"],
        date_value=data.get("edit_date_value"),
    )


def _format_admin_reschedule_date_label(target_date: datetime.date) -> str:
    weekdays = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]
    return f"{target_date.strftime('%d.%m')} · {weekdays[target_date.weekday()]}"


async def _build_admin_reschedule_date_options(duration: int) -> list[tuple[str, str]]:
    salon_now = get_salon_now()
    booking_window = max(int(salon_config.get("booking_window", 7) or 7), 1)
    working_days = salon_config.get("working_days", [1, 2, 3, 4, 5, 6, 0])
    blacklisted_dates = set(salon_config.get("blacklisted_dates", []))
    options: list[tuple[str, str]] = []

    for offset in range(booking_window):
        target_date = salon_now.date() + timedelta(days=offset)
        date_value = target_date.strftime("%d.%m.%Y")
        js_weekday = (target_date.weekday() + 1) % 7
        if js_weekday not in working_days or date_value in blacklisted_dates:
            continue

        times = await _build_admin_reschedule_time_options(date_value, duration)
        if times:
            options.append((_format_admin_reschedule_date_label(target_date), date_value))

    return options


async def _build_admin_reschedule_time_options(
    date_value: str,
    duration: int,
    *,
    current_date: str | None = None,
    current_time: str | None = None,
) -> list[str]:
    start_mins, end_mins = parse_working_hours(salon_config.get("working_hours", "10:00-20:00"))
    interval = int(salon_config.get("schedule_interval", 30) or 30)
    if interval <= 0:
        interval = 30

    busy_slots = await database.get_busy_slots_by_date(date_value)
    if current_date == date_value and current_time:
        busy_slots = [
            busy
            for busy in busy_slots
            if not (busy.get("time") == current_time and int(busy.get("duration") or 60) == int(duration or 60))
        ]

    salon_now = get_salon_now()
    salon_today = salon_now.strftime("%d.%m.%Y")
    current_salon_mins = salon_now.hour * 60 + salon_now.minute if date_value == salon_today else -1

    available_times: list[str] = []
    for slot_start in range(start_mins, end_mins - int(duration or 60) + 1, interval):
        if slot_start <= current_salon_mins:
            continue

        is_busy = False
        for busy in busy_slots:
            try:
                busy_h, busy_m = map(int, busy["time"].split(":"))
            except (KeyError, ValueError, AttributeError):
                continue
            busy_start = busy_h * 60 + busy_m
            busy_duration = int(busy.get("duration") or 60)
            if slot_overlaps(slot_start, int(duration or 60), busy_start, busy_duration):
                is_busy = True
                break

        if not is_busy:
            available_times.append(f"{slot_start // 60:02d}:{slot_start % 60:02d}")

    return available_times


def _build_admin_reschedule_dates_markup(booking_id: int, options: list[tuple[str, str]]):
    from aiogram.utils.keyboard import InlineKeyboardBuilder

    builder = InlineKeyboardBuilder()
    for label, date_value in options:
        builder.row(types.InlineKeyboardButton(text=label, callback_data=f"admin_resched_date|{booking_id}|{date_value}"))
    builder.row(types.InlineKeyboardButton(text="❌ Отмена", callback_data=f"admin_resched_cancel|{booking_id}"))
    return builder.as_markup()


def _build_admin_reschedule_times_markup(booking_id: int, date_value: str, times: list[str]):
    from aiogram.utils.keyboard import InlineKeyboardBuilder

    builder = InlineKeyboardBuilder()
    row: list[types.InlineKeyboardButton] = []
    for time_value in times:
        row.append(types.InlineKeyboardButton(text=time_value, callback_data=f"admin_resched_time|{booking_id}|{date_value}|{time_value}"))
        if len(row) == 3:
            builder.row(*row)
            row = []
    if row:
        builder.row(*row)
    builder.row(types.InlineKeyboardButton(text="❌ Отмена", callback_data=f"admin_resched_cancel|{booking_id}"))
    return builder.as_markup()


def _build_admin_reschedule_confirm_markup(booking_id: int, date_value: str, time_value: str):
    from aiogram.utils.keyboard import InlineKeyboardBuilder

    builder = InlineKeyboardBuilder()
    builder.row(types.InlineKeyboardButton(text="✅ Подтвердить", callback_data=f"admin_resched_confirm|{booking_id}|{date_value}|{time_value}"))
    builder.row(types.InlineKeyboardButton(text="❌ Отмена", callback_data=f"admin_resched_cancel|{booking_id}"))
    return builder.as_markup()


@router.callback_query(F.data.startswith("booking_reschedule|"))
async def admin_booking_reschedule_start_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    _, booking_id_str, context, date_token, source_filter, page_str = callback.data.split("|", 5)
    booking = await database.get_booking_record_by_id(int(booking_id_str))
    if not booking:
        await callback.answer("Запись не найдена.", show_alert=True)
        return

    _id, user_id, name, phone, current_date, current_time, status, duration = booking
    if status != "scheduled":
        await callback.answer("Эту запись уже нельзя перенести.", show_alert=True)
        return

    date_options = await _build_admin_reschedule_date_options(int(duration or 60))
    if not date_options:
        await callback.answer("Свободных дат для переноса сейчас нет.", show_alert=True)
        return

    await state.set_state(AdminRescheduleBookingForm.waiting_for_date)
    await state.update_data(
        booking_id=int(booking_id_str),
        booking_context=context,
        booking_date_value=_normalize_context_value(date_token),
        booking_source_filter=source_filter,
        booking_page=int(page_str),
        booking_name=name,
        booking_phone=phone,
        booking_user_id=user_id,
        current_date=current_date,
        current_time=current_time,
        duration=int(duration or 60),
    )
    await callback.message.answer(
        (
            "🔁 <b>Перенос записи</b>\n\n"
            f"<b>Сейчас:</b> {escape(current_date)} в {escape(current_time)}\n"
            "Выберите новую дату:"
        ),
        parse_mode="HTML",
        reply_markup=_build_admin_reschedule_dates_markup(int(booking_id_str), date_options),
    )


@router.callback_query(F.data.startswith("admin_resched_date|"))
async def admin_booking_reschedule_date_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    _, booking_id_str, date_value = callback.data.split("|", 2)
    data = await state.get_data()
    if data.get("booking_id") != int(booking_id_str):
        await callback.answer("Сессия переноса устарела. Начните заново.", show_alert=True)
        return

    time_options = await _build_admin_reschedule_time_options(
        date_value,
        int(data["duration"]),
        current_date=data.get("current_date"),
        current_time=data.get("current_time"),
    )
    if not time_options:
        await callback.answer("На эту дату нет свободного времени.", show_alert=True)
        return

    await state.set_state(AdminRescheduleBookingForm.waiting_for_time)
    await state.update_data(new_date=date_value)
    await callback.message.edit_text(
        (
            "🔁 <b>Перенос записи</b>\n\n"
            f"<b>Новая дата:</b> {escape(date_value)}\n"
            "Выберите новое время:"
        ),
        parse_mode="HTML",
        reply_markup=_build_admin_reschedule_times_markup(int(booking_id_str), date_value, time_options),
    )


@router.callback_query(F.data.startswith("admin_resched_time|"))
async def admin_booking_reschedule_time_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    _, booking_id_str, date_value, time_value = callback.data.split("|", 3)
    data = await state.get_data()
    if data.get("booking_id") != int(booking_id_str):
        await callback.answer("Сессия переноса устарела. Начните заново.", show_alert=True)
        return

    await state.set_state(AdminRescheduleBookingForm.waiting_for_confirmation)
    await state.update_data(new_date=date_value, new_time=time_value)
    await callback.message.edit_text(
        (
            "🔁 <b>Подтвердите перенос</b>\n\n"
            f"<b>Было:</b> {escape(data['current_date'])} в {escape(data['current_time'])}\n"
            f"<b>Станет:</b> {escape(date_value)} в {escape(time_value)}"
        ),
        parse_mode="HTML",
        reply_markup=_build_admin_reschedule_confirm_markup(int(booking_id_str), date_value, time_value),
    )


@router.callback_query(F.data.startswith("admin_resched_confirm|"))
async def admin_booking_reschedule_confirm_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await callback.answer()
    _, booking_id_str, date_value, time_value = callback.data.split("|", 3)
    data = await state.get_data()
    if data.get("booking_id") != int(booking_id_str):
        await callback.answer("Сессия переноса устарела. Начните заново.", show_alert=True)
        return

    moved = await database.reschedule_booking_if_available(int(booking_id_str), date_value, time_value)
    if not moved:
        await callback.answer("Этот слот уже занят. Выберите другое время.", show_alert=True)
        return

    await _notify_client_about_admin_booking_action(
        callback.bot,
        user_id=data.get("booking_user_id"),
        title="🔁 <b>Запись перенесена</b>",
        lines=[
            f"<b>Услуга:</b> {escape(data.get('service_name') or '—')}",
            f"<b>Было:</b> {escape(data['current_date'])} в {escape(data['current_time'])}",
            f"<b>Стало:</b> {escape(date_value)} в {escape(time_value)}",
            "Пожалуйста, сохраните новое время.",
        ],
    )

    await state.clear()
    await callback.message.edit_text(
        (
            "✅ <b>Запись перенесена</b>\n\n"
            f"<b>Новая дата:</b> {escape(date_value)}\n"
            f"<b>Новое время:</b> {escape(time_value)}"
        ),
        parse_mode="HTML",
    )
    await callback.message.answer(
        (
            "📝 <b>Карточка записи обновлена</b>\n\n"
            f"<b>Клиент:</b> {escape(data['booking_name'])}\n"
            f"<b>Телефон:</b> {escape(data['booking_phone'])}\n"
            f"<b>Было:</b> {escape(data['current_date'])} в {escape(data['current_time'])}\n"
            f"<b>Стало:</b> {escape(date_value)} в {escape(time_value)}"
        ),
        parse_mode="HTML",
    )
    await _show_booking_card(
        callback.message,
        int(booking_id_str),
        context=data["booking_context"],
        page=int(data["booking_page"]),
        source_filter=data["booking_source_filter"],
        date_value=data.get("booking_date_value"),
    )


@router.callback_query(F.data.startswith("admin_resched_cancel|"))
async def admin_booking_reschedule_cancel_callback(callback: types.CallbackQuery, state):
    admin_id = getenv("ADMIN_ID")
    if not admin_id or str(callback.from_user.id) != admin_id:
        return

    await state.clear()
    await callback.answer("Перенос отменен")
    await callback.message.edit_text(
        "❌ <b>Перенос отменен</b>\n\nЕсли нужно, можно открыть перенос заново из карточки записи.",
        parse_mode="HTML",
    )


